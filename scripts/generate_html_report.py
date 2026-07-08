"""Generate a self-contained, editable HTML report of PCCT qualification gate findings.

Reads gate_results/, tracker/, and the workflow + case-summary spreadsheets.
Writes gate_results/qualification_report.html (self-contained, all images embedded).

The HTML has text marked contenteditable="true" so it can be edited live in
the browser. A "Download edited HTML" button serializes the current DOM
back to a file. Numeric data blocks (gate_summary.txt content) are kept
non-editable.

Run from PCCT/ root or from PCCT/scripts/:
    python scripts/generate_html_report.py
"""
import base64
import csv
import io
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "gate_results" / "qualification_report.html"


def b64_image(p: Path) -> str:
    if not p.exists():
        return ""
    return "data:image/png;base64," + base64.b64encode(p.read_bytes()).decode("ascii")


def fig_to_b64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight")
    plt.close(fig)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def read(p: Path) -> str:
    return p.read_text(encoding="utf-8") if p.exists() else ""


def extract_block(text: str, header_re: str, end_re: str = r"^={5,}") -> str:
    m = re.search(header_re, text, re.MULTILINE)
    if not m:
        return ""
    rest = text[m.end():]
    end = re.search(end_re, rest, re.MULTILINE)
    return rest[: end.start()].strip() if end else rest.strip()


def html_escape(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ── Plot builders ─────────────────────────────────────────────────────────────

def load_snr_csv(p: Path):
    """Return dict pid -> {mean_hu, std, snr} (skip rows with err)."""
    out = {}
    for r in csv.DictReader(open(p, encoding="utf-8")):
        if r.get("err"):
            continue
        try:
            pid = r["case"].split("_Bv_")[0]
            out[pid] = {
                "mean_hu": float(r["mean"]),
                "std": float(r["std"]),
                "snr": float(r["snr"]),
            }
        except (ValueError, KeyError):
            pass
    return out


def plot_contrast_timing(pcct, eid, threshold=250):
    """Paired-bar plot: per-patient aortic mean HU for PCCT and EID, threshold line."""
    paired_pids = sorted(set(pcct) & set(eid))
    pcct_vals = [pcct[p]["mean_hu"] for p in paired_pids]
    eid_vals = [eid[p]["mean_hu"] for p in paired_pids]

    fig, ax = plt.subplots(figsize=(11, 4.5))
    x = np.arange(len(paired_pids))
    w = 0.38
    ax.bar(x - w/2, pcct_vals, w, label=f"PCCT (N={len(pcct_vals)})", color="#2563eb", edgecolor="white")
    ax.bar(x + w/2, eid_vals, w, label=f"EID (N={len(eid_vals)})", color="#f59e0b", edgecolor="white")
    ax.axhline(threshold, color="#dc2626", linestyle="--", linewidth=1.2,
               label=f"Threshold {threshold} HU")
    ax.set_xticks(x)
    ax.set_xticklabels(paired_pids, rotation=60, ha="right", fontsize=8)
    ax.set_ylabel("Mean aortic HU (10 mm ROI at centroid)")
    ax.set_title("Gate 1.2 Contrast Timing — paired PCCT vs EID aortic enhancement")
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y", alpha=0.25)
    ax.set_ylim(bottom=0)
    return fig


def plot_snr_paired(pcct, eid):
    """Side-by-side: paired bar chart of SNR, and scatter."""
    paired_pids = sorted(set(pcct) & set(eid))
    p = [pcct[i]["snr"] for i in paired_pids]
    e = [eid[i]["snr"] for i in paired_pids]
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 4.5), gridspec_kw={"width_ratios": [2.4, 1]})

    # Left: paired bars
    x = np.arange(len(paired_pids)); w = 0.38
    ax1.bar(x - w/2, p, w, label="PCCT", color="#2563eb", edgecolor="white")
    ax1.bar(x + w/2, e, w, label="EID", color="#f59e0b", edgecolor="white")
    ax1.set_xticks(x); ax1.set_xticklabels(paired_pids, rotation=60, ha="right", fontsize=8)
    ax1.set_ylabel("SNR (mean / std HU)")
    ax1.set_title("Gate 1.3 Image Noise — paired SNR (aortic ROI)")
    ax1.legend(fontsize=9); ax1.grid(axis="y", alpha=0.25)

    # Right: scatter with y=x
    mx = max(max(p), max(e)) * 1.1
    ax2.plot([0, mx], [0, mx], "k--", alpha=0.4, linewidth=1, label="y=x (parity)")
    ax2.scatter(e, p, c="#2563eb", s=44, zorder=5, edgecolor="white")
    ax2.set_xlabel("EID SNR")
    ax2.set_ylabel("PCCT SNR")
    ax2.set_title(f"PCCT vs EID SNR\nratio {np.mean([a/b for a,b in zip(p,e)]):.2f}× (PCCT/EID)")
    ax2.set_xlim(0, mx); ax2.set_ylim(0, mx); ax2.set_aspect("equal")
    ax2.legend(fontsize=8); ax2.grid(alpha=0.25)

    fig.tight_layout()
    return fig


def load_effort():
    """Return (paired_pids, pcct_eff, ccta_eff) for patients with numeric effort on both."""
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "PCCT_CCTA_Case_Summaries.xlsx", data_only=True)
    ws = wb["Case Summaries"]
    hdr = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    eff_col = hdr.index("Analyst Effort\n(1=Most, 5=Least)") + 1

    by_pid = defaultdict(dict)
    for r in range(3, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        scan = ws.cell(r, 2).value
        eff = ws.cell(r, eff_col).value
        if not pid or scan not in ("PCCT", "CCTA"):
            continue
        try:
            e = float(eff)
        except (TypeError, ValueError):
            continue
        by_pid[pid][scan] = e

    pids = sorted([pid for pid, d in by_pid.items() if "PCCT" in d and "CCTA" in d])
    return pids, [by_pid[p]["PCCT"] for p in pids], [by_pid[p]["CCTA"] for p in pids]


def plot_effort(pids, pcct_eff, ccta_eff):
    """Two-panel: paired effort bars, and histogram of deltas."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 4.5), gridspec_kw={"width_ratios": [2.4, 1]})

    x = np.arange(len(pids)); w = 0.38
    ax1.bar(x - w/2, pcct_eff, w, label="PCCT", color="#2563eb", edgecolor="white")
    ax1.bar(x + w/2, ccta_eff, w, label="CCTA", color="#f59e0b", edgecolor="white")
    ax1.set_xticks(x); ax1.set_xticklabels(pids, rotation=60, ha="right", fontsize=8)
    ax1.set_ylabel("Analyst effort (1=most, 5=least)")
    ax1.set_title(f"Gate 2.4 Analyst Editing Effort — paired (N={len(pids)})")
    ax1.set_yticks([1, 2, 3, 4, 5])
    ax1.set_ylim(0, 5.5)
    ax1.legend(fontsize=9); ax1.grid(axis="y", alpha=0.25)

    deltas = [p - c for p, c in zip(pcct_eff, ccta_eff)]
    mean_d = np.mean(deltas)
    bins = np.arange(-4.25, 4.25, 0.5)
    ax2.hist(deltas, bins=bins, color="#2563eb", edgecolor="white", alpha=0.8)
    ax2.axvline(0, color="black", linewidth=1, alpha=0.5)
    ax2.axvline(mean_d, color="#dc2626", linewidth=1.5, label=f"mean {mean_d:+.2f}")
    ax2.set_xlabel("Effort delta (PCCT − CCTA)\n← PCCT harder    PCCT easier →")
    ax2.set_ylabel("Patients")
    ax2.set_title(f"Effort delta distribution\nmedian {np.median(deltas):+.1f}")
    ax2.legend(fontsize=9); ax2.grid(alpha=0.25)

    fig.tight_layout()
    return fig


# ── Main ──────────────────────────────────────────────────────────────────────

def config_matrix_html():
    """Build a cross-configuration Gate 3 verdict matrix (canonical region) from
    gate_results/variants/, so the report presents all configs equally rather than
    a single headline. Primary metric: log-wCV for process outputs, untransformed
    for plaque."""
    vdir = ROOT / "gate_results" / "variants"
    configs = [
        ("legacy rms-rel", "summary_A_legacy-rmsrel.txt"),
        ("var-comp (default)", "summary_B_variance-component_DEFAULT.txt"),
        ("var-comp + scanner-term", "summary_C_variance-component_scanner-term.txt"),
        ("var-comp + OQ bias", "summary_D_oq-bias-criterion.txt"),
    ]
    endpoints = [("Lumen Volume", "log"), ("Wall Volume", "log"), ("Vessel Volume", "log"),
                 ("CALC Volume", "unt"), ("LRNC Volume", "unt"),
                 ("NonCALC Matrix Volume", "unt"), ("Total Plaque Volume", "unt")]
    hdr_ep = re.compile(r"--- (.+?) \(mm.*\[length-normalized\] ---")
    hdr_any = re.compile(r"^\s*---.*---\s*$")

    def parse_canonical(path):
        if not path.exists():
            return {}
        txt = path.read_text(encoding="utf-8").splitlines()
        b = [i for i, l in enumerate(txt) if "SUB-SEGMENT INTERSECTION ANALYSIS" in l]
        lines = txt[:b[0]] if b else txt
        cur, blocks = None, {}
        for l in lines:
            m = hdr_ep.match(l.strip())
            if m:
                cur = m.group(1).strip(); blocks.setdefault(cur, {"log": None, "unt": None, "v": None})
            elif hdr_any.match(l):
                cur = None
            if cur:
                lm = re.search(r"Log-wCV:\s+([\d.]+)%", l); um = re.search(r"Untransformed wCV:\s+([\d.]+)%", l)
                vm = re.search(r"Overlap:\s+.*?(PASS|FAIL)", l)
                if lm and blocks[cur]["log"] is None: blocks[cur]["log"] = lm.group(1)
                if um and blocks[cur]["unt"] is None: blocks[cur]["unt"] = um.group(1)
                if vm and blocks[cur]["v"] is None: blocks[cur]["v"] = vm.group(1)
        return blocks
    data = {name: parse_canonical(vdir / fn) for name, fn in configs}

    h = ['<table><tr><th>Endpoint (primary metric)</th>']
    for name, _ in configs:
        h.append(f"<th>{name}</th>")
    h.append("</tr>\n")
    for ep, metric in endpoints:
        h.append(f"<tr><td>{ep} ({'log' if metric=='log' else 'untransf.'})</td>")
        for name, _ in configs:
            b = data.get(name, {}).get(ep)
            if not b or b[metric] is None:
                h.append("<td>—</td>")
            else:
                cls = "pass" if b["v"] == "PASS" else "fail"
                h.append(f'<td>{b[metric]}% <span class="{cls}">{b["v"] or "?"}</span></td>')
        h.append("</tr>\n")
    h.append("</table>")
    return "".join(h)


def comparison_table_html(kind):
    """Render gate_results/{gate3,gate4}_comparison.csv as an OQ-vs-PCCT side-by-side
    table with 95% CIs."""
    path = ROOT / "gate_results" / f"{kind}_comparison.csv"
    if not path.exists():
        return "<p class='meta'><em>(comparison table not generated — re-run run_gate_analyses.py)</em></p>"
    rows = list(csv.DictReader(open(path, encoding="utf-8")))

    def verdict(v):
        if v == "YES":
            return '<span class="pass">overlap ✓</span>'
        if v == "NO":
            return '<span class="fail">no overlap ✗</span>'
        return v

    if kind == "gate3":
        h = ['<table><tr><th>Endpoint</th><th>Metric</th>'
             '<th>OQ delta wCV [95% CI]</th><th>PCCT wCV [95% CI]</th><th>CI overlap</th></tr>']
        for r in rows:
            h.append(f"<tr><td>{r['endpoint']}</td><td>{r['metric']}</td>"
                     f"<td>{r['oq_wcv']}% [{r['oq_ci_lo']}, {r['oq_ci_hi']}]</td>"
                     f"<td>{r['pcct_wcv']}% [{r['pcct_ci_lo']}, {r['pcct_ci_hi']}]</td>"
                     f"<td>{verdict(r['ci_overlap'])}</td></tr>")
        h.append("</table>")
        return "".join(h)
    # gate4
    h = ['<table><tr><th>Endpoint (log scale)</th><th>OQ bias [95% CI]</th><th>OQ LoA</th>'
         '<th>PCCT bias [95% CI]</th><th>PCCT LoA</th><th>bias CI overlap</th></tr>']
    for r in rows:
        h.append(f"<tr><td>{r['endpoint']}</td>"
                 f"<td>{r['oq_bias']} [{r['oq_bias_ci_lo']}, {r['oq_bias_ci_hi']}]</td>"
                 f"<td>[{r['oq_loa_lo']}, {r['oq_loa_hi']}]</td>"
                 f"<td>{r['pcct_bias']} [{r['pcct_bias_ci_lo']}, {r['pcct_bias_ci_hi']}]</td>"
                 f"<td>[{r['pcct_loa_lo']}, {r['pcct_loa_hi']}]</td>"
                 f"<td>{verdict(r['bias_ci_overlap'])}</td></tr>")
    h.append("</table>")
    return "".join(h)


def main():
    gate_summary = read(ROOT / "gate_results" / "gate_summary.txt")
    gate1 = extract_block(gate_summary, r"^GATE 1.*?$")
    gate2 = extract_block(gate_summary, r"^GATE 2.*?$")
    gate3 = extract_block(gate_summary, r"^GATE 3.*?$")
    gate4 = extract_block(gate_summary, r"^GATE 4 .{,2}SYSTEMATIC", end_re=r"^GATE 4 SUPPLEMENTARY")

    # BA plots (canonical only)
    plot_dir = ROOT / "gate_results" / "bland_altman_plots"
    ba_vars = ["LumenVol", "WallVol", "VesselVol", "CALCVol", "LRNCVol",
               "NonCALCMATXVol", "TotalPlaqueVolume"]
    ba_canonical = {v: b64_image(plot_dir / f"BA_{v}.png") for v in ba_vars}

    # New Gate 1 / Gate 2.4 plots — generate from raw data
    snr_pcct = load_snr_csv(ROOT / "gate_results" / "snr_pcct.csv")
    snr_eid = load_snr_csv(ROOT / "gate_results" / "snr_eid.csv")
    contrast_img = fig_to_b64(plot_contrast_timing(snr_pcct, snr_eid))
    snr_img = fig_to_b64(plot_snr_paired(snr_pcct, snr_eid))

    eff_pids, eff_pcct, eff_ccta = load_effort()
    effort_img = fig_to_b64(plot_effort(eff_pids, eff_pcct, eff_ccta))

    # Audit + case reviews
    audit_md = read(ROOT / "tracker" / "workitem-audit.md")
    m = re.search(r"## Summary\s*\n\n(.*?)\n\n", audit_md, re.DOTALL)
    audit_summary_md = m.group(1) if m else ""
    case_reviews_md = read(ROOT / "tracker" / "case-reviews.md")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>PCCT Qualification - Gate Findings Report</title>
<style>
:root {{
    --bg: #fafbfc; --panel: #fff; --border: #d1d5db; --text: #1f2937;
    --muted: #6b7280; --pass: #16a34a; --fail: #dc2626; --warn: #f59e0b; --accent: #2563eb;
}}
body {{ background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; line-height: 1.55; max-width: 1100px; margin: 0 auto; padding: 24px; padding-top: 64px; }}
h1, h2, h3, h4 {{ color: var(--text); margin-top: 1.4em; }}
h1 {{ font-size: 1.85em; border-bottom: 2px solid var(--accent); padding-bottom: 8px; }}
h2 {{ font-size: 1.4em; border-bottom: 1px solid var(--border); padding-bottom: 4px; }}
h3 {{ font-size: 1.15em; color: var(--accent); }}
.meta {{ color: var(--muted); font-size: 0.9em; }}
.panel {{ background: var(--panel); border: 1px solid var(--border); border-radius: 6px; padding: 16px 20px; margin: 12px 0; }}
.pass {{ color: var(--pass); font-weight: 600; }}
.fail {{ color: var(--fail); font-weight: 600; }}
.warn {{ color: var(--warn); font-weight: 600; }}
table {{ border-collapse: collapse; width: 100%; margin: 8px 0 16px; font-size: 0.92em; }}
th {{ background: #f3f4f6; text-align: left; padding: 8px 10px; border-bottom: 2px solid var(--border); }}
td {{ padding: 6px 10px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }}
tr:hover {{ background: #f9fafb; }}
pre {{ background: #f3f4f6; padding: 12px; border-radius: 4px; overflow-x: auto; font-size: 0.85em; font-family: Consolas, Menlo, monospace; white-space: pre-wrap; }}
.callout {{ border-left: 4px solid var(--accent); background: #eff6ff; padding: 10px 14px; margin: 10px 0; }}
.callout-warn {{ border-left-color: var(--warn); background: #fffbeb; }}
img.fig {{ max-width: 100%; height: auto; border: 1px solid var(--border); border-radius: 4px; margin: 6px 0; }}
.ba-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(360px, 1fr)); gap: 12px; }}
.ba-grid figure {{ margin: 0; }}
.ba-grid figcaption {{ font-size: 0.85em; color: var(--muted); margin-top: 4px; }}
#editbar {{ position: fixed; top: 0; left: 0; right: 0; background: #1f2937; color: white; padding: 10px 20px; z-index: 1000; box-shadow: 0 2px 4px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 12px; }}
#editbar button {{ background: var(--accent); color: white; border: 0; padding: 6px 14px; border-radius: 4px; cursor: pointer; font-size: 0.92em; }}
#editbar button:hover {{ background: #1d4ed8; }}
#editbar .info {{ font-size: 0.85em; opacity: 0.85; }}
[contenteditable="true"]:focus {{ outline: 2px solid var(--accent); outline-offset: 2px; background: #fefce8; }}
[contenteditable="true"]:hover {{ background: #fef9c3; }}
.uneditable {{ background: #f3f4f6; padding: 8px 12px; border-radius: 4px; font-size: 0.85em; color: var(--muted); border-left: 3px solid var(--muted); }}
</style>
</head>
<body>

<div id="editbar">
    <strong>PCCT Qualification Report</strong>
    <span class="info">Yellow-highlighted text is editable — click to edit</span>
    <span style="flex:1"></span>
    <button onclick="downloadEdited()">Download edited HTML</button>
    <button onclick="window.print()">Print / PDF</button>
</div>

<h1 contenteditable="true">PCCT Qualification — Gate Findings Report <span style="font-size:0.6em;color:var(--accent)">· Version 2 (current)</span></h1>
<p class="meta" contenteditable="true">Generated {now} from <code>gate_results/gate_summary.txt</code> (variance-component wCV, default config), <code>gate_results/variants/</code>, <code>tracker/*.md</code>, paired SNR, and case-summary spreadsheet. Reference: B.1P Delta Validation OQ (730-CVV-040 v0.1) and Original B.1P OQ (4-B1P-033 v2.0). The original analysis is frozen in <code>gate_results_v1_original/</code>.</p>

<section>
<h2>Executive Summary</h2>
<div class="callout callout-warn" contenteditable="true">
<strong>No single headline verdict.</strong> Gate 3/4 outcomes depend on the estimator and region choice, which are still being finalized — the configurations are presented side by side below rather than as one result. Methodology: <code>tracker/statistical-methodology.md</code>. <strong>N = 25 paired patients</strong>, preliminary (target ≥ 30); PT-124 excluded (no overlapping vessel). Latest 2026-07-07 workitem data.
</div>
<table>
<tr><th>Gate</th><th>Requirement</th><th>Status (see config matrix)</th></tr>
<tr><td><strong>Gate 1</strong> — Technical & Image Quality</td><td>DICOM, contrast timing, SNR, kernel</td><td><span class="pass">PASS</span></td></tr>
<tr><td><strong>Gate 2</strong> — Workflow Integration</td><td>Ingestion, centerline, lumen/wall editing</td><td><span class="warn">REVIEW</span></td></tr>
<tr><td><strong>Gate 3</strong> — Reproducibility (wCV)</td><td>95% CI overlap with delta OQ</td><td>config-dependent — process outputs pass only with scanner-term or on sub-segment region</td></tr>
<tr><td><strong>Gate 4</strong> — Bias & Agreement</td><td>BA bias vs OQ</td><td><span class="fail">NonCALC Matrix & Total Plaque bias FAIL</span> (both criteria)</td></tr>
</table>
</section>

<section>
<h2>Configuration Variants — Gate 3 wCV (canonical region, N=25)</h2>
<p class="meta" contenteditable="true">Primary-metric within-subject CV and 95%-CI-overlap verdict vs the delta OQ, under each estimator/criterion. Full per-config reports in <code>gate_results/variants/</code>; corrected-estimator rationale in <code>tracker/statistical-methodology.md</code>. Sub-segment region (extent-matched) generally passes but is on stale segmentations pending regeneration.</p>
{config_matrix_html()}
<div class="callout" contenteditable="true">
<strong>Reading it:</strong> legacy→variance-component corrects the wCV (esp. the log branch), flipping Lumen/Vessel PASS→FAIL on the canonical region; the scanner term (removing systematic modality bias) brings process outputs back within OQ. Plaque wCV passes broadly; the binding issue is plaque <em>bias</em> (Gate 4).
</div>
</section>

<section>
<h2>Methodology Notes</h2>
<div class="callout" contenteditable="true">
<strong>Vessel-overlap normalization (canonical as of 2026-05-08).</strong> For each PCCT/EID pair, summaries are summed only over the (bodySite, vessel-location) intersection — so the same anatomical extent is measured on each scanner. Patients with no overlapping vessel are excluded (PT-124). Most pairs are partial-overlap; PCCT typically traces 1-3 additional distal vessels (PDA, PLB, marginals, diagonals).
</div>
<div class="callout" contenteditable="true">
<strong>Metric per endpoint family.</strong> Log-wCV (and log-scale Bland-Altman) for process outputs — Lumen, Wall, Vessel Volume — matching 4-B1P-033 Table 7 and 730-CVV-040 Table 6. Untransformed wCV (and untransformed BA) for plaque volumes (CALC, LRNC, NonCALC Matrix, Total Plaque), matching 4-B1P-033 Table 9 reporting scale.
</div>
<div class="callout callout-warn" contenteditable="true">
<strong>Wall bias threshold is project-specific.</strong> |bias| &lt; 10% of mean was added 2026-05-08 in response to case-review finding (PT-142) of likely true modality bias for Wall at high disease. Not derivable from 4-B1P-033 or 730-CVV-040 (both assess Wall reproducibility by wCV only). Threshold matched to plaque-component bias thresholds since Wall segmentation is dominated by the wall-plaque boundary.
</div>
</section>

<section>
<h2>Gate 1 — Technical & Image Quality Prerequisites</h2>

<h3>1.2 Contrast Timing — paired aortic enhancement</h3>
<img class="fig" src="{contrast_img}" alt="Contrast timing paired bars">
<p class="meta" contenteditable="true">Threshold: peak aortic HU ≥ 250 in ≥ 90% of cases. PCCT 28/28 (100%) PASS. EID 24/25 (96%) PASS. Only PT-124 (242 HU) below threshold.</p>

<h3>1.3 Image Noise — paired SNR</h3>
<img class="fig" src="{snr_img}" alt="SNR paired and scatter">
<p class="meta" contenteditable="true">SNR = mean_HU / std_HU within a 10 mm aortic ROI at centroid. Threshold: PCCT SNR non-inferior to EID (ratio ≥ 0.85). PCCT mean SNR ~2× EID; lower noise in 24/25 paired cases. PT-142 and PT-158 EID excluded — Aorta.nrrd not generated.</p>

<pre class="uneditable">{html_escape(gate1)}</pre>
<div class="panel" contenteditable="true">
<strong>Reviewer commentary:</strong> Editable space for Gate 1 notes.
</div>
</section>

<section>
<h2>Gate 2 — Workflow Integration</h2>
<pre class="uneditable">{html_escape(gate2)}</pre>

<h3>2.4 Lumen & Wall Editing — analyst effort comparison</h3>
<img class="fig" src="{effort_img}" alt="Effort paired bars and delta histogram">
<p class="meta" contenteditable="true">Per-case analyst effort (1=most effort, 5=least), recorded in PCCT_CCTA_Case_Summaries.xlsx. Paired N={len(eff_pids)} patients with numeric scores on both scans. PCCT mean {np.mean(eff_pcct):.2f}, CCTA mean {np.mean(eff_ccta):.2f}; delta +{np.mean(eff_pcct)-np.mean(eff_ccta):.2f} (median {np.median([p-c for p,c in zip(eff_pcct, eff_ccta)]):.1f}). Symmetric distribution → PCCT does not impose meaningfully higher editing effort.</p>

<div class="panel" contenteditable="true">
<strong>Reviewer commentary:</strong> Centerline coverage and editing-effort findings.
</div>
</section>

<section>
<h2>Gate 3 — Quantitative Reproducibility</h2>
<h3>OQ reference vs PCCT result — wCV with 95% CIs</h3>
<p class="meta" contenteditable="true">Primary metric per endpoint (log-wCV for process outputs, untransformed for plaque), variance-component estimator, canonical region, N=25. Acceptance = 95% CI overlap with the delta OQ.</p>
{comparison_table_html("gate3")}
<pre class="uneditable">{html_escape(gate3)}</pre>
<div class="panel" contenteditable="true">
<strong>Reviewer commentary:</strong> The block above is the <em>variance-component (default)</em> config. Under the corrected (OQ-consistent) estimator the process outputs (Lumen/Wall/Vessel) FAIL the CI-overlap on the canonical region — the systematic modality bias inflates the raw cross-scanner wCV. They come within OQ with the <em>scanner-term</em> (bias removed) or on the <em>sub-segment</em> region (extent matched). See the configuration matrix above and <code>tracker/statistical-methodology.md</code>. Plaque wCV passes broadly.
</div>
</section>

<section>
<h2>Gate 4 — Bias & Agreement</h2>
<h3>OQ reference vs PCCT result — plaque Bland-Altman bias (log scale) with 95% CIs</h3>
<p class="meta" contenteditable="true">Plaque BA on log(x+1) scale (matches 730-CVV-040 Table 6; raw plaque volumes are heteroscedastic). Acceptance = PCCT bias 95% CI overlaps the OQ Table 6 bias 95% CI. The BA plots below overlay the OQ bias line, OQ bias CI, and OQ LoA band so overlap is visible.</p>
{comparison_table_html("gate4")}
<pre class="uneditable">{html_escape(gate4)}</pre>

<h3>Bland-Altman plots</h3>
<div class="ba-grid">
"""
    for v in ba_vars:
        if ba_canonical[v]:
            html += f'<figure><img class="fig" src="{ba_canonical[v]}" alt="BA {v}"><figcaption contenteditable="true">{v}</figcaption></figure>\n'
    html += """</div>

<div class="panel" contenteditable="true">
<strong>Reviewer commentary (latest 07-07 data):</strong> The Gate 4 bias criterion has two forms — the legacy project-specific <code>|bias|&lt;5%/10% of mean</code> and the OQ-consistent <code>oq-ci-overlap</code> (PCCT log-scale BA bias 95% CI vs 730-CVV-040 Table 6). <strong>NonCALC Matrix and Total Plaque bias FAIL under both</strong> — a real, statistically distinguishable modality bias (PCCT systematically lower). Wall bias has grown to ~28% across re-processing. LRNC "passes" the CI-overlap test only by low power (its CI is too wide to reject). Full detail in <code>gate_results/variants/</code> and <code>tracker/statistical-methodology.md §2.1</code>.
</div>
</section>

<section>
<h2>Case Reviews</h2>
<div class="panel" contenteditable="true">
"""
    cr = case_reviews_md
    cr = re.sub(r"^# (.*)$", r"<h3>\1</h3>", cr, flags=re.MULTILINE)
    cr = re.sub(r"^### (.*)$", r"<h4>\1</h4>", cr, flags=re.MULTILINE)
    cr = re.sub(r"^---$", r"<hr>", cr, flags=re.MULTILINE)
    cr = re.sub(r"^\- \*\*(.+?):\*\* (.*)$", r"<p><strong>\1:</strong> \2</p>", cr, flags=re.MULTILINE)
    cr = re.sub(r"^\- (.*)$", r"<li>\1</li>", cr, flags=re.MULTILINE)
    cr = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", cr)
    cr = re.sub(r"`(.+?)`", r"<code>\1</code>", cr)
    html += cr
    html += """</div>
</section>

<section>
<h2>Workitem Audit Summary</h2>
<div class="panel" contenteditable="true">
<p>From <code>tracker/workitem-audit.md</code> — audits all 81 workitems in the workflow spreadsheet against the ip3-manager instance and local <code>workitem_summaries/</code> state.</p>
"""
    audit_lines = audit_summary_md.split("\n")
    if audit_lines:
        html += "<table>\n"
        for line in audit_lines:
            if not line.strip() or line.strip().startswith("|---"):
                continue
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            tag = "th" if cells[0] == "Status" else "td"
            html += "<tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>\n"
        html += "</table>\n"
    html += """<p><em>Carolyn Walker's 13 MISSING_LOCAL workitems are intentionally not in the gate-analysis cohort.</em></p>
</div>
</section>

<script>
function downloadEdited() {
    const clone = document.documentElement.cloneNode(true);
    const bar = clone.querySelector('#editbar');
    if (bar) bar.remove();
    const html = '<!DOCTYPE html>\\n' + clone.outerHTML;
    const blob = new Blob([html], {type: 'text/html'});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const ts = new Date().toISOString().replace(/[:.]/g, '-').slice(0,16);
    a.download = `qualification_report_edited_${ts}.html`;
    a.click();
    URL.revokeObjectURL(url);
}
</script>
</body>
</html>
"""

    OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"  size: {len(html.encode('utf-8'))//1024} KB")


if __name__ == "__main__":
    main()
