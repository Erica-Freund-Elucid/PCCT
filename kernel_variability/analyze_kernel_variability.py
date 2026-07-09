"""Convolution-kernel + auto-segmentation variability characterization (single subject).

Input: a "combined" workitem summary workbook holding one PCCT acquisition of one
patient reconstructed with many convolution kernels and AUTO-segmented (no reader
edits), so the only sources of plaque variability are (a) the convolution kernel
and (b) the auto-segmentation algorithm.

The workbook's "PatientLevel" sheet has one row per (workitem, target); this script
aggregates to patient level by summing across targets per workitem (= per kernel).

It separates the two variability sources using the exact identity
    log(Volume) = log(Volume/Len) + log(Len)
so raw-volume variance splits into a per-mm "density" term and a traced-"extent"
(Len) term, and it reports extent-invariant composition ratios (component fractions,
plaque burden) where the traced extent cancels — isolating the pure kernel effect.

Outputs (written next to this script, in outputs/):
    per_kernel_patient_level.csv   one row per kernel: totals, per-mm, ratios, factors
    variability_summary.csv        raw / per-mm / extent variability per endpoint
    variance_decomposition.csv     log-identity split (extent vs density) + correlation
    composition_ratios.csv         extent-invariant ratios: CV% + Qr-vs-Bv family effect

Usage:
    python kernel_variability/analyze_kernel_variability.py \
        ["<path to combined workitem summary .xlsx>"]
Default input: PT-142_workitem_summary_combined.xlsx in the repo root.
"""
import csv
import os
import re
import sys
from collections import defaultdict

import numpy as np
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.join(ROOT, "PT-142_workitem_summary_combined.xlsx")
OUT_DIR = os.path.join(HERE, "outputs")

# Composition/structure volumes we aggregate per kernel (patient-level totals).
VOLS = ["CALCVol", "LRNCVol", "NonCALCMATXVol", "TotalPlaqueVolume",
        "LumenVol", "WallVol", "LumenAndWallVol", "Len"]


def load_per_kernel(xlsx_path):
    """Return {workitemID: {kernel, sums...}} aggregated across targets."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["PatientLevel"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ci = {h: i for i, h in enumerate(hdr)}
    agg = defaultdict(lambda: defaultdict(float))
    kernel = {}
    patient = None
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        wid = row[ci["workitemID"]]
        if not wid:
            continue
        patient = row[ci["individualID"]]
        kernel[wid] = str(row[ci["convolutionKernel"]])
        for v in VOLS:
            val = row[ci[v]]
            if isinstance(val, (int, float)):
                agg[wid][v] += val
    return patient, kernel, agg


def parse_kernel(k):
    """Qr40u_2 -> (family='Qr', sharpness=40, mode='u', strength=2)."""
    m = re.match(r"(Qr|Bv)(\d+)([ud])_(\d)", k)
    if not m:
        return (k, None, None, None)
    return (m.group(1), int(m.group(2)), m.group(3), int(m.group(4)))


def cv(a):
    a = np.asarray(a, float)
    return float(a.std(ddof=1) / a.mean() * 100) if a.mean() else 0.0


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.isfile(xlsx):
        sys.exit(f"Input not found: {xlsx}")
    os.makedirs(OUT_DIR, exist_ok=True)
    patient, kernel, agg = load_per_kernel(xlsx)
    W = list(agg)
    fac = {w: parse_kernel(kernel[w]) for w in W}
    fam = {w: fac[w][0] for w in W}
    n = len(W)
    print(f"Patient {patient}: {n} auto reconstructions "
          f"(Qr={sum(f=='Qr' for f in fam.values())}, Bv={sum(f=='Bv' for f in fam.values())})")

    # ---- per-kernel patient-level table ----
    ratios = {
        "PlaqueBurden_TP_over_LW": lambda a: a["TotalPlaqueVolume"] / a["LumenAndWallVol"],
        "WallToLumen": lambda a: a["WallVol"] / a["LumenVol"],
        "CALCfrac_of_plaque": lambda a: a["CALCVol"] / a["TotalPlaqueVolume"],
        "NonCALCfrac_of_plaque": lambda a: a["NonCALCMATXVol"] / a["TotalPlaqueVolume"],
        "CALC_over_Wall": lambda a: a["CALCVol"] / a["WallVol"],
        "NonCALCMATX_over_Wall": lambda a: a["NonCALCMATXVol"] / a["WallVol"],
    }
    with open(os.path.join(OUT_DIR, "per_kernel_patient_level.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["workitemID", "kernel", "family", "sharpness", "mode", "strength"]
                   + VOLS + [f"{v}_per_mm" for v in VOLS if v != "Len"] + list(ratios))
        for wid in W:
            a = agg[wid]
            fa = fac[wid]
            row = [wid, kernel[wid], fa[0], fa[1], fa[2], fa[3]]
            row += [f"{a[v]:.4f}" for v in VOLS]
            row += [f"{a[v] / a['Len']:.6f}" for v in VOLS if v != "Len"]
            row += [f"{fn(a):.6f}" for fn in ratios.values()]
            w.writerow(row)

    # ---- variability summary (raw totals, per-mm, extent) ----
    with open(os.path.join(OUT_DIR, "variability_summary.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["endpoint", "raw_mean", "raw_CV%", "raw_min", "raw_max",
                    "raw_range%mean", "per_mm_mean", "per_mm_CV%"])
        for v in VOLS:
            raw = np.array([agg[x][v] for x in W])
            rmean = raw.mean()
            line = [v, f"{rmean:.3f}", f"{cv(raw):.1f}", f"{raw.min():.3f}",
                    f"{raw.max():.3f}", f"{(raw.max() - raw.min()) / rmean * 100:.1f}"]
            if v != "Len":
                pm = np.array([agg[x][v] / agg[x]["Len"] for x in W])
                line += [f"{pm.mean():.6f}", f"{cv(pm):.1f}"]
            else:
                line += ["", ""]
            w.writerow(line)

    # ---- variance decomposition: log V = log(density) + log Len ----
    Ln = np.array([agg[x]["Len"] for x in W])
    with open(os.path.join(OUT_DIR, "variance_decomposition.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["endpoint", "SD_logV_total%", "SD_logLen_extent%",
                    "SD_logDensity_permm%", "corr_Len_density"])
        for v in ["TotalPlaqueVolume", "CALCVol", "NonCALCMATXVol", "WallVol", "LumenVol"]:
            V = np.array([agg[x][v] for x in W])
            lv, lL, ld = np.log(V), np.log(Ln), np.log(V / Ln)
            r = float(np.corrcoef(lL, ld)[0, 1])
            w.writerow([v, f"{lv.std(ddof=1) * 100:.1f}", f"{lL.std(ddof=1) * 100:.1f}",
                        f"{ld.std(ddof=1) * 100:.1f}", f"{r:.2f}"])

    # ---- extent-invariant composition ratios: CV% + family (Qr vs Bv) effect ----
    with open(os.path.join(OUT_DIR, "composition_ratios.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ratio", "mean", "CV%", "min", "max", "Bv_mean", "Qr_mean", "Qr_minus_Bv_%"])
        for name, fn in ratios.items():
            a = np.array([fn(agg[x]) for x in W])
            bv = np.mean([fn(agg[x]) for x in W if fam[x] == "Bv"])
            qr = np.mean([fn(agg[x]) for x in W if fam[x] == "Qr"])
            w.writerow([name, f"{a.mean():.4f}", f"{cv(a):.1f}", f"{a.min():.4f}",
                        f"{a.max():.4f}", f"{bv:.4f}", f"{qr:.4f}",
                        f"{100 * (qr - bv) / a.mean():+.0f}"])

    print(f"Wrote 4 CSVs to {os.path.relpath(OUT_DIR, ROOT)}/")


if __name__ == "__main__":
    main()
