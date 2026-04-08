"""Generate PCCT Scanner Qualification Tracker workbook.

Reads tracker/*.md files to pick up any Owner/Status/Evidence/Notes
that have already been filled in, then writes the Excel workbook.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# --- Parse markdown tracker files ---

def parse_tracker_md(filepath):
    """Parse a gate tracker markdown file and return a dict keyed by criterion ID."""
    results = {}
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
    except FileNotFoundError:
        return results

    # Split into sections by ### headers
    sections = re.split(r"^### ", content, flags=re.MULTILINE)
    for section in sections[1:]:  # skip preamble
        lines = section.strip().splitlines()
        # Extract ID from header like "1.4 Reconstruction Kernel" or "A.1 HU..."
        header_match = re.match(r"(\w+\.\d+)\s", lines[0])
        if not header_match:
            continue
        cid = header_match.group(1)

        fields = {}
        for line in lines:
            for field in ["Owner", "Status", "Evidence", "Notes"]:
                m = re.match(rf"^-\s+\*\*{field}:\*\*\s*(.*)", line)
                if m:
                    fields[field.lower()] = m.group(1).strip()
        results[cid] = fields
    return results


TRACKER_FILES = {
    "gate1": r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\tracker\gate1-technical-prerequisites.md",
    "gate2": r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\tracker\gate2-workflow-integration.md",
    "gate3": r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\tracker\gate3-reproducibility.md",
    "gate4": r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\tracker\gate4-bias-agreement.md",
    "advisory": r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\tracker\advisory-operational-checks.md",
}

md_data = {}
for key, path in TRACKER_FILES.items():
    md_data.update(parse_tracker_md(path))

# Map some status variations
STATUS_MAP = {
    "complete": "Pass",
    "completed": "Pass",
    "done": "Pass",
    "pass": "Pass",
    "fail": "Fail",
    "failed": "Fail",
    "in progress": "In progress",
    "not started": "Not started",
    "blocked": "Blocked",
}


def get_md_fields(cid):
    """Get owner, status, evidence, notes from parsed markdown."""
    fields = md_data.get(cid, {})
    owner = fields.get("owner", "")
    status_raw = fields.get("status", "Not started").lower().strip()
    status = STATUS_MAP.get(status_raw, fields.get("status", "Not started"))
    evidence = fields.get("evidence", "")
    notes = fields.get("notes", "")
    return owner, status, evidence, notes


# --- Excel setup ---
wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
gate_header_font = Font(bold=True, size=14, color="2F5496")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

status_validation = DataValidation(
    type="list",
    formula1='"Not started,In progress,Pass,Fail,Blocked"',
    allow_blank=True,
)
status_validation.error = "Please select a valid status"
status_validation.errorTitle = "Invalid Status"

COLUMNS = [
    ("ID", 6),
    ("Criterion", 28),
    ("Threshold", 32),
    ("Method", 32),
    ("Owner", 16),
    ("Status", 14),
    ("Evidence Link / Location", 30),
    ("Evidence Description", 28),
    ("Date Completed", 16),
    ("Notes", 36),
]

GATES = {
    "Gate 1 - Technical": {
        "title": "Gate 1 — Technical & Image Quality Prerequisites",
        "subtitle": "Required — Scanner not eligible without passing all items",
        "criteria": [
            ("1.1", "DICOM Compliance", "Full DICOM 3.0; mandatory tags present", "Automated tag validation on ingestion"),
            ("1.2", "Contrast Timing", "Peak aortic HU ≥ 300 in ≥ 90% of cases", "Aortic ROI measurement across paired dataset"),
            ("1.3", "Image Noise (SNR/CNR)", "Non-inferior to reference CTA (±15%)", "Uniform phantom or aortic ROI SD comparison"),
            ("1.4", "Reconstruction Kernel", "Soft-tissue kernel available; sharp kernel documented", "Protocol review + DICOM header"),
        ],
    },
    "Gate 2 - Workflow": {
        "title": "Gate 2 — Workflow Integration Checks",
        "subtitle": "Workflow step — Engineering remediation required before performance testing",
        "criteria": [
            ("2.1", "DICOM Ingestion", "100% successful ingestion; zero WIID assignment failures", "Run full paired dataset through ingest pipeline"),
            ("2.2", "Centerline / Vessel Tree", "Extraction success rate ≥ 95%; major vessel coverage equivalent to reference", "Compare vessel segment count and length distribution"),
            ("2.3", "Lumen & Wall Initialization", "Auto-initialization without manual override in ≥ 85% of segments", "Log override frequency per scanner type"),
            ("2.4", "Lumen & Wall Editing", "Edit rate not significantly higher than reference CTA (95% CI overlap)", "Track editing events per case; include cases Carolyn already completed"),
            ("2.5", "Plaque Quantification", "See Gate 3 — core quantitative criteria", "Paired CTA vs. PCCT comparison"),
            ("2.6", "Report Generation", "All required fields populated; no missing-data warnings", "ECR review in physician viewer and PDF reports for N ≥ 10 cases; use ones Carolyn already did"),
        ],
    },
    "Gate 3 - Reproducibility": {
        "title": "Gate 3 — Core Quantitative Reproducibility (Paired CTA vs. PCCT)",
        "subtitle": "Performance — wCV from un-edited output, manual centerline for N ≥ 30. All thresholds must pass.",
        "criteria": [
            ("3.1", "Lumen Volume & Vessel Length", "≤ 10% wCV (non-inferiority)", "wCV = (SD of differences / mean) × 100; report 95% CI"),
            ("3.2", "Calcified Plaque Volume", "≤ 20% wCV", "wCV = (SD of differences / mean) × 100; report 95% CI"),
            ("3.3", "Wall Volume", "≤ 30% wCV", "wCV = (SD of differences / mean) × 100; report 95% CI"),
            ("3.4", "Total Plaque Volume", "≤ 30% wCV", "wCV = (SD of differences / mean) × 100; report 95% CI"),
        ],
    },
    "Gate 4 - Bias": {
        "title": "Gate 4 — Systematic Bias & Agreement",
        "subtitle": "Performance — Flag for clinical review on failure; may restrict use to specific indications",
        "criteria": [
            ("4.1", "Mean Bias — Lumen Volume", "|bias| < 5% of mean lumen volume", "Bland-Altman plot; report LoA"),
            ("4.2", "Mean Bias — Calcified Plaque", "|bias| < 10% of mean calc volume", "Bland-Altman plot; proportional bias test"),
            ("4.3", "Limits of Agreement", "LoA within ±1.96 SD; no proportional bias (r² < 0.1)", "Regress residuals on mean"),
        ],
    },
    "Advisory": {
        "title": "Advisory — Contextual & Operational Checks",
        "subtitle": "Inform labelling and use restrictions; do not block qualification",
        "criteria": [
            ("A.1", "HU Calibration Stability", "Verify 130 HU threshold valid; map spectral to conventional HU if needed", "Scanner calibration review"),
            ("A.2", "Radiation Dose", "Document DLP/CTDI; no dose increase that offsets clinical utility", "Protocol dose comparison"),
            ("A.3", "Patient Subgroup Stratification", "Report wCV by calcium burden and BMI", "Subgroup heterogeneity analysis"),
            ("A.4", "Reader/Operator Variability", "Inter-reader editing variability within expected range", "Second analyst completes N = 10 Carolyn already did"),
            ("A.5", "Compatibility with Ongoing Projects", "No new failure modes with AVTE/AWAL", "Regression testing"),
            ("A.6", "Software Version Lock", "Qualify against plaque algos ≥ Nov 2025; re-qualify on major updates", "Version documentation"),
        ],
    },
}


def build_gate_sheet(ws, gate):
    ws.sheet_properties.tabColor = "2F5496"
    ws.add_data_validation(status_validation)

    ws.merge_cells("A1:J1")
    ws["A1"] = gate["title"]
    ws["A1"].font = gate_header_font

    ws.merge_cells("A2:J2")
    ws["A2"] = gate["subtitle"]
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for row_offset, (cid, name, threshold, method) in enumerate(gate["criteria"]):
        row = 5 + row_offset

        # Pull data from markdown if available
        md_owner, md_status, md_evidence, md_notes = get_md_fields(cid)

        values = [
            cid, name, threshold, method,
            md_owner,           # Owner
            md_status,          # Status
            md_evidence,        # Evidence Link
            "",                 # Evidence Description
            "",                 # Date Completed
            md_notes,           # Notes
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = wrap
            cell.border = thin_border
            if col_idx == 6:
                cell.alignment = wrap_center

        status_validation.add(ws.cell(row=row, column=6))

    last_row = 4 + len(gate["criteria"])
    status_range = f"F5:F{last_row}"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Pass"'], fill=green_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Fail"'], fill=red_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"In progress"'], fill=yellow_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Blocked"'], fill=red_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Not started"'], fill=gray_fill))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{last_row}"


first = True
for sheet_name, gate in GATES.items():
    if first:
        ws = wb.active
        ws.title = sheet_name
        first = False
    else:
        ws = wb.create_sheet(title=sheet_name)
    build_gate_sheet(ws, gate)

# --- Summary Dashboard ---
ds = wb.create_sheet(title="Summary", index=0)
ds.sheet_properties.tabColor = "1F4E79"

ds.merge_cells("A1:F1")
ds["A1"] = "PCCT Scanner Qualification — Summary Dashboard"
ds["A1"].font = Font(bold=True, size=16, color="2F5496")

ds.merge_cells("A3:F3")
ds["A3"] = "Fill in the tracker sheets. This summary auto-updates from status columns."
ds["A3"].font = Font(italic=True, size=10, color="666666")

summary_headers = ["Gate", "Total Criteria", "Pass", "Fail", "In Progress", "Gate Result"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ds.cell(row=5, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_center
    cell.border = thin_border

col_widths = [28, 16, 10, 10, 14, 16]
for i, w in enumerate(col_widths, 1):
    ds.column_dimensions[get_column_letter(i)].width = w

gate_info = [
    ("Gate 1 — Technical Prerequisites", "'Gate 1 - Technical'", 4),
    ("Gate 2 — Workflow Integration", "'Gate 2 - Workflow'", 6),
    ("Gate 3 — Reproducibility", "'Gate 3 - Reproducibility'", 4),
    ("Gate 4 — Bias & Agreement", "'Gate 4 - Bias'", 3),
    ("Advisory — Operational Checks", "'Advisory'", 6),
]

for row_offset, (label, sheet_ref, count) in enumerate(gate_info):
    row = 6 + row_offset
    ds.cell(row=row, column=1, value=label).border = thin_border
    ds.cell(row=row, column=1).alignment = wrap
    ds.cell(row=row, column=2, value=count).border = thin_border
    ds.cell(row=row, column=2).alignment = wrap_center

    rng = f"{sheet_ref}!F5:F{4 + count}"
    ds.cell(row=row, column=3).border = thin_border
    ds.cell(row=row, column=3).alignment = wrap_center
    ds.cell(row=row, column=3, value=f'=COUNTIF({rng},"Pass")')

    ds.cell(row=row, column=4).border = thin_border
    ds.cell(row=row, column=4).alignment = wrap_center
    ds.cell(row=row, column=4, value=f'=COUNTIF({rng},"Fail")')

    ds.cell(row=row, column=5).border = thin_border
    ds.cell(row=row, column=5).alignment = wrap_center
    ds.cell(row=row, column=5, value=f'=COUNTIF({rng},"In progress")')

    result_cell = ds.cell(row=row, column=6)
    result_cell.border = thin_border
    result_cell.alignment = wrap_center
    result_cell.value = f'=IF(C{row}=B{row},"PASS",IF(D{row}>0,"FAIL","PENDING"))'

result_range = "F6:F10"
ds.conditional_formatting.add(result_range, CellIsRule(operator="equal", formula=['"PASS"'], fill=green_fill))
ds.conditional_formatting.add(result_range, CellIsRule(operator="equal", formula=['"FAIL"'], fill=red_fill))
ds.conditional_formatting.add(result_range, CellIsRule(operator="equal", formula=['"PENDING"'], fill=yellow_fill))

ds.freeze_panes = "A6"

output = r"C:\Users\EricaFreund\OneDrive - Elucid Bioimaging Inc\PCCT\PCCT_Qualification_Tracker.xlsx"
wb.save(output)
print(f"Saved to {output}")
print("\nData pulled from markdown files:")
for cid, fields in sorted(md_data.items()):
    if any(v for v in fields.values()):
        print(f"  {cid}: {fields}")
